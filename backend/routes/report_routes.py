from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from auth import get_current_user
import models
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def check_permission(user):
    if user.role not in [models.RoleEnum.admin, models.RoleEnum.supervisor]:
        raise HTTPException(status_code=403, detail="Not authorized")

@router.get("/complaints/excel")
def download_complaints_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    check_permission(current_user)
    complaints = db.query(models.Complaint).order_by(
        models.Complaint.created_at.desc()
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints Report"

    # Header style
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        "Complaint No.", "Title", "Customer", "Category",
        "Priority", "Status", "Assigned To", "Created Date", "Updated Date"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    status_colors = {
        "Open":       "DBEAFE", "Assigned":    "EDE9FE",
        "In Progress":"FEF3C7", "Escalated":   "FEE2E2",
        "Resolved":   "D1FAE5", "Closed":      "F3F4F6",
    }
    priority_colors = {
        "Critical": "450a0a", "High":   "FEE2E2",
        "Medium":   "FEF3C7", "Low":    "D1FAE5",
    }

    for row, c in enumerate(complaints, 2):
        status   = c.status.value   if hasattr(c.status,   'value') else str(c.status)
        priority = c.priority.value if hasattr(c.priority, 'value') else str(c.priority)
        values = [
            c.complaint_number,
            c.title,
            c.customer.name  if c.customer  else "—",
            c.category.name  if c.category  else "—",
            priority,
            status,
            c.agent.name     if c.agent     else "Unassigned",
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "—",
            c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "—",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 5:  # Priority
                fill_color = priority_colors.get(priority, "F3F4F6")
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if col == 6:  # Status
                fill_color = status_colors.get(status, "F3F4F6")
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["CCRTS Report Summary"])
    ws2.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws2.append([])
    ws2.append(["Total Complaints", len(complaints)])
    for status_val in ["Open", "Assigned", "In Progress", "Escalated", "Resolved", "Closed"]:
        count = sum(
            1 for c in complaints
            if (c.status.value if hasattr(c.status, 'value') else str(c.status)) == status_val
        )
        ws2.append([status_val, count])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"CCRTS_Complaints_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/complaints/pdf")
def download_complaints_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    check_permission(current_user)
    complaints = db.query(models.Complaint).order_by(
        models.Complaint.created_at.desc()
    ).limit(100).all()

    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("CCRTS — Complaints Report", styles['Title']))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total: {len(complaints)} complaints",
        styles['Normal']
    ))
    elements.append(Spacer(1, 12))

    # Table data
    data = [["Complaint No.", "Title", "Customer", "Priority", "Status", "Category", "Date"]]
    for c in complaints:
        status   = c.status.value   if hasattr(c.status,   'value') else str(c.status)
        priority = c.priority.value if hasattr(c.priority, 'value') else str(c.priority)
        title    = c.title[:40] + "..." if len(c.title) > 40 else c.title
        data.append([
            c.complaint_number,
            title,
            c.customer.name if c.customer else "—",
            priority,
            status,
            c.category.name if c.category else "—",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "—",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 9),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE',    (0,1), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',     (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    filename = f"CCRTS_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )